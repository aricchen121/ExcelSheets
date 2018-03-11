import openpyxl

#load workbook for using
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']


print(sheet.cell(row=1, column=2))

print(sheet.cell(row=1, column=2).value)

#print rows and the items
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
